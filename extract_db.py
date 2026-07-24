import openpyxl
import json
import os

excel_path = r"C:\Users\pla_y\.gemini\antigravity\scratch\Celestia\Planetary.xlsx"
output_dir = r"C:\Users\pla_y\.gemini\antigravity\scratch\Celestia"
output_path = os.path.join(output_dir, "database.js")

# Ensure output directory exists
os.makedirs(output_dir, exist_ok=True)

print("Loading workbook...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Planetary']

print("Parsing rows...")
rows = []
for r in range(2, sheet.max_row + 1):
    eq = sheet.cell(row=r, column=3).value
    fa_en = sheet.cell(row=r, column=4).value
    fa_th = sheet.cell(row=r, column=5).value
    fb_en = sheet.cell(row=r, column=6).value
    fb_th = sheet.cell(row=r, column=7).value
    fc_en = sheet.cell(row=r, column=8).value
    fc_th = sheet.cell(row=r, column=9).value
    desc_en = sheet.cell(row=r, column=10).value
    desc_th = sheet.cell(row=r, column=11).value
    conf = sheet.cell(row=r, column=12).value

    # We skip empty rows or rows without an equation
    if not eq:
        continue

    rows.append({
        'eq': str(eq).strip(),
        'a_en': str(fa_en).strip() if fa_en else '',
        'a_th': str(fa_th).strip() if fa_th else '',
        'b_en': str(fb_en).strip() if fb_en else '',
        'b_th': str(fb_th).strip() if fb_th else '',
        'c_en': str(fc_en).strip() if fc_en else '',
        'c_th': str(fc_th).strip() if fc_th else '',
        'desc_en': str(desc_en).strip() if desc_en else '',
        'desc_th': str(desc_th).strip() if desc_th else '',
        'conf': str(conf).strip() if conf else ''
    })

print(f"Total parsed rows: {len(rows)}")

# Write to database.js
print(f"Writing to {output_path}...")
with open(output_path, 'w', encoding='utf-8') as f:
    f.write("// Planetary Web App Database\n")
    f.write("// Auto-generated from Planetary.xlsx\n\n")
    f.write("const PLANETARY_DB = ")
    json.dump(rows, f, ensure_ascii=False, indent=2)
    f.write(";\n")

print("Done extraction!")
