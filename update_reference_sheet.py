import openpyxl

excel_path = r'C:\Users\pla_y\.gemini\antigravity\scratch\Celestia\Planetary.xlsx'
wb = openpyxl.load_workbook(excel_path)
ws = wb['Reference']

# 1. Update row 19 (House-Ruler-Combination description)
ws.cell(19, 2).value = 'ตารางคู่ผสม เรือนที่เจ้าเรือนถูกจร (X) x เรือนที่กำลังเคลื่อนเข้า (N) — 144 แถว พร้อมมุมตีความหลายด้าน 5 Bullets ภาษาไทยที่เป็นรูปธรรม เข้าใจง่าย สำหรับ UI แบบเลือก dropdown'
ws.cell(19, 4).value = 'สังเคราะห์จาก Sheet House-Ruler-Transit (อ้างอิง Brady) และเกลาเรียบเรียงภาษาใหม่ให้กระชับ เหมาะกับแอป'

# 2. Append missing sheets to Reference table (Rows 22 to 26)
new_reference_rows = [
    [
        'Aspect-Principles',
        'หลักการ ทฤษฎี และตารางค่าออร์บ (Orb) ของมุมสัมพันธ์ (Hard vs Soft, Element/Mode Compatibility, Harmonic Division)',
        '"Aspects in Astrology" โดย Sue Tompkins (Element Books / Destiny Books, 1989/2002)',
        'แปลและสรุปเป็นภาษาไทย พร้อมตารางออร์บมาตรฐานตัวอย่าง',
        'ระบุแหล่งอ้างอิงและเลขหน้าต้นฉบับรายแถวในคอลัมน์ "แหล่งที่มา" ชัดเจน'
    ],
    [
        'Aspect-Natal-Meaning',
        'ความหมายของมุมสัมพันธ์แต่ละประเภทในดวงกำเนิด (กุม, เล็ง, สแควร์, ตรีโกณ, เซกส์ไทล์ ฯลฯ)',
        '"Aspects in Astrology" โดย Sue Tompkins',
        'แปลและสรุปแก่นความหมาย (Keywords) และคำอธิบายภาษาไทย-อังกฤษ',
        'ระบุแหล่งอ้างอิงรายแถวในคอลัมน์ "แหล่งที่มา"'
    ],
    [
        'Aspect-Natal-Cookbook',
        'คำแปลและคู่มือตีความคู่ดาวทำมุมสัมพันธ์กันในดวงกำเนิด/ดวงจร (Aspect Combinations รายคู่)',
        '"Aspects in Astrology" โดย Sue Tompkins',
        'แปลและเรียบเรียงคู่ดาวทำมุมสัมพันธ์กัน พร้อมคำอธิบาย TH-EN',
        'ระบุแหล่งอ้างอิงและเลขหน้าต้นฉบับรายแถว'
    ],
    [
        'Aspect-Natal-Angles',
        'ความหมายเมื่อดาวทำมุมสัมพันธ์กับแกนสำคัญของดวงชะตา (Ascendant, Descendant, MC, IC)',
        '"Aspects in Astrology" โดย Sue Tompkins',
        'แปลและจัดหมวดหมู่แยกตามแกนทั้ง 4 จุดพร้อมคำอธิบาย',
        'ระบุแหล่งอ้างอิงรายแถวในคอลัมน์ "แหล่งที่มา"'
    ],
    [
        'Planet-Sign',
        'ความหมายสังเคราะห์ของดาวเคราะห์เมื่อสถิตในแต่ละราศี (12 ราศี x รายการดาวและจุดสำคัญ)',
        '"Horoscope Symbols" โดย Robert Hand ผสมผสานกับ "The Inner Sky" โดย Steven Forrest',
        'สังเคราะห์จาก Keyword ของ Robert Hand และ Steven Forrest พร้อมคำนวณสถานะมิตรธาตุ (Dignity)',
        'ระบุแหล่งอ้างอิงและเลขหน้าต้นฉบับแยกตามผู้แต่งในคอลัมน์ "แหล่งอ้างอิง"'
    ]
]

start_row = 22
for i, r_data in enumerate(new_reference_rows):
    curr_row = start_row + i
    for col_idx, val in enumerate(r_data, 1):
        cell = ws.cell(curr_row, col_idx)
        cell.value = val

wb.save(excel_path)
print(f'Successfully updated Reference sheet in Planetary.xlsx! Total rows in Reference table: {ws.max_row}')
