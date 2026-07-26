import openpyxl
import json
import os

excel_path = r"C:\Users\pla_y\.gemini\antigravity\scratch\Celestia\Planetary.xlsx"
db_extra_path = r"C:\Users\pla_y\.gemini\antigravity\scratch\Celestia\database_extra.js"

print("Loading workbook...")
wb = openpyxl.load_workbook(excel_path, data_only=True)

# 1. Parse Rulership sheet
sheet_rule = wb['Rulership']
rulership_rows = []
for r in range(2, sheet_rule.max_row + 1):
    id_val = sheet_rule.cell(row=r, column=1).value
    sign_th = sheet_rule.cell(row=r, column=2).value
    sign_en = sheet_rule.cell(row=r, column=3).value
    trad_th = sheet_rule.cell(row=r, column=4).value
    trad_en = sheet_rule.cell(row=r, column=5).value
    mod_th = sheet_rule.cell(row=r, column=6).value
    mod_en = sheet_rule.cell(row=r, column=7).value
    source = sheet_rule.cell(row=r, column=8).value
    notes = sheet_rule.cell(row=r, column=9).value

    if not sign_th:
        continue

    rulership_rows.append({
        "id": id_val,
        "sign_th": str(sign_th).strip() if sign_th else "",
        "sign_en": str(sign_en).strip() if sign_en else "",
        "trad_th": str(trad_th).strip() if trad_th else "",
        "trad_en": str(trad_en).strip() if trad_en else "",
        "mod_th": str(mod_th).strip() if mod_th else "",
        "mod_en": str(mod_en).strip() if mod_en else "",
        "source": str(source).strip() if source else "",
        "notes": str(notes).strip() if notes else ""
    })

print(f"Parsed {len(rulership_rows)} Rulership rows.")

# 2. Parse House-Ruler-Transit sheet
sheet_lord = wb['House-Ruler-Transit']
lord_rows = []
for r in range(2, sheet_lord.max_row + 1):
    id_val = sheet_lord.cell(row=r, column=1).value
    house_num = sheet_lord.cell(row=r, column=2).value
    area_th = sheet_lord.cell(row=r, column=3).value
    meaning_th = sheet_lord.cell(row=r, column=4).value
    page = sheet_lord.cell(row=r, column=5).value
    source = sheet_lord.cell(row=r, column=6).value
    conf = sheet_lord.cell(row=r, column=7).value

    if not house_num:
        continue

    lord_rows.append({
        "id": id_val,
        "house": int(house_num),
        "area_th": str(area_th).strip() if area_th else "",
        "meaning_th": str(meaning_th).strip() if meaning_th else "",
        "page": str(page).strip() if page else "",
        "source": str(source).strip() if source else "",
        "conf": str(conf).strip() if conf else ""
    })

print(f"Parsed {len(lord_rows)} House-Ruler-Transit rows.")

# 3. Parse House-Ruler-Combination sheet
sheet_comb = wb['House-Ruler-Combination']
comb_rows = []
for r in range(2, sheet_comb.max_row + 1):
    id_val = sheet_comb.cell(row=r, column=1).value
    hx = sheet_comb.cell(row=r, column=2).value
    hn = sheet_comb.cell(row=r, column=3).value
    area_x = sheet_comb.cell(row=r, column=4).value
    area_n = sheet_comb.cell(row=r, column=5).value
    synth_th = sheet_comb.cell(row=r, column=6).value
    source = sheet_comb.cell(row=r, column=7).value
    conf = sheet_comb.cell(row=r, column=8).value
    notes = sheet_comb.cell(row=r, column=9).value
    bullets_raw = sheet_comb.cell(row=r, column=10).value

    if not hx or not hn:
        continue

    bullets_list = [b.strip() for b in bullets_raw.split(" || ") if b.strip()] if bullets_raw else []

    comb_rows.append({
        "id": id_val,
        "hx": int(hx),
        "hn": int(hn),
        "area_x": str(area_x).strip() if area_x else "",
        "area_n": str(area_n).strip() if area_n else "",
        "synth_th": str(synth_th).strip() if synth_th else "",
        "bullets": bullets_list,
        "source": str(source).strip() if source else "",
        "conf": str(conf).strip() if conf else "",
        "notes": str(notes).strip() if notes else ""
    })

print(f"Parsed {len(comb_rows)} House-Ruler-Combination rows.")

# 4. Parse Planet-Key-Principle sheet
sheet_pkp = wb['Planet-Key-Principle']
pkp_rows = []
for r in range(2, sheet_pkp.max_row + 1):
    id_val = sheet_pkp.cell(row=r, column=1).value
    planet_th = sheet_pkp.cell(row=r, column=2).value
    planet_en = sheet_pkp.cell(row=r, column=3).value
    key_principle_en = sheet_pkp.cell(row=r, column=4).value
    nature_th = sheet_pkp.cell(row=r, column=5).value
    style_th = sheet_pkp.cell(row=r, column=6).value
    page = sheet_pkp.cell(row=r, column=7).value
    source = sheet_pkp.cell(row=r, column=8).value

    if not planet_th:
        continue

    pkp_rows.append({
        "id": id_val,
        "planet_th": str(planet_th).strip() if planet_th else "",
        "planet_en": str(planet_en).strip() if planet_en else "",
        "key_principle_en": str(key_principle_en).strip() if key_principle_en else "",
        "nature_th": str(nature_th).strip() if nature_th else "",
        "style_th": str(style_th).strip() if style_th else "",
        "page": str(page).strip() if page else "",
        "source": str(source).strip() if source else ""
    })

print(f"Parsed {len(pkp_rows)} Planet-Key-Principle rows.")

# 5. Parse Planet-House-Ruler-Matrix sheet
sheet_phrm = wb['Planet-House-Ruler-Matrix']
phrm_rows = []
for r in range(2, sheet_phrm.max_row + 1):
    id_val = sheet_phrm.cell(row=r, column=1).value
    planet_th = sheet_phrm.cell(row=r, column=2).value
    planet_en = sheet_phrm.cell(row=r, column=3).value
    hx = sheet_phrm.cell(row=r, column=4).value
    area_x = sheet_phrm.cell(row=r, column=5).value
    nature = sheet_phrm.cell(row=r, column=6).value
    synth_th = sheet_phrm.cell(row=r, column=7).value
    page = sheet_phrm.cell(row=r, column=8).value
    source = sheet_phrm.cell(row=r, column=9).value
    conf = sheet_phrm.cell(row=r, column=10).value
    notes = sheet_phrm.cell(row=r, column=11).value

    if not planet_th or not hx:
        continue

    phrm_rows.append({
        "id": id_val,
        "planet_th": str(planet_th).strip() if planet_th else "",
        "planet_en": str(planet_en).strip() if planet_en else "",
        "hx": int(hx),
        "area_x": str(area_x).strip() if area_x else "",
        "nature": str(nature).strip() if nature else "",
        "synth_th": str(synth_th).strip() if synth_th else "",
        "page": str(page).strip() if page else "",
        "source": str(source).strip() if source else "",
        "conf": str(conf).strip() if conf else "",
        "notes": str(notes).strip() if notes else ""
    })

print(f"Parsed {len(phrm_rows)} Planet-House-Ruler-Matrix rows.")

# Read existing database_extra.js up to RULERSHIP_DB
with open(db_extra_path, 'r', encoding='utf-8') as f:
    content = f.read()

if "const RULERSHIP_DB =" in content:
    content = content.split("const RULERSHIP_DB =")[0].strip()

js_addition = "\n\n// Rulership & House Ruler Transit Databases (Auto-generated from Planetary.xlsx)\n"
js_addition += "const RULERSHIP_DB = " + json.dumps(rulership_rows, ensure_ascii=False, indent=2) + ";\n\n"
js_addition += "const HOUSE_RULER_TRANSIT_DB = " + json.dumps(lord_rows, ensure_ascii=False, indent=2) + ";\n\n"
js_addition += "const HOUSE_RULER_COMB_DB = " + json.dumps(comb_rows, ensure_ascii=False, indent=2) + ";\n\n"
js_addition += "const PLANET_KEY_PRINCIPLE_DB = " + json.dumps(pkp_rows, ensure_ascii=False, indent=2) + ";\n\n"
js_addition += "const PLANET_HOUSE_RULER_DB = " + json.dumps(phrm_rows, ensure_ascii=False, indent=2) + ";\n"

new_content = content.strip() + js_addition

with open(db_extra_path, 'w', encoding='utf-8') as f:
    f.write(new_content)

print("Successfully updated database_extra.js with all 5 Lord databases!")
