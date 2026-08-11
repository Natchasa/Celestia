import openpyxl
import json
import os

excel_path = r"C:\Users\pla_y\.gemini\antigravity\scratch\Celestia\Planetary.xlsx"
# NOTE: output_dir points at non-github/ on purpose — database_extra.js holds
# the real proprietary astrology data and must never land in the top-level
# folder that gets pushed to the public GitHub repo. See non-github/database_extra.js.
output_dir = r"C:\Users\pla_y\.gemini\antigravity\scratch\Celestia\non-github"
db_extra_path = os.path.join(output_dir, "database_extra.js")

print("Loading workbook...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Astrocartography']

print("Parsing rows...")
rows = []
for r in range(2, sheet.max_row + 1):
    category = sheet.cell(row=r, column=2).value
    topic_en = sheet.cell(row=r, column=3).value
    topic_th = sheet.cell(row=r, column=4).value
    planet_en = sheet.cell(row=r, column=5).value
    planet_th = sheet.cell(row=r, column=6).value
    keyword = sheet.cell(row=r, column=7).value
    desc_th = sheet.cell(row=r, column=8).value
    desc_en = sheet.cell(row=r, column=9).value
    page = sheet.cell(row=r, column=10).value
    source = sheet.cell(row=r, column=11).value
    confidence = sheet.cell(row=r, column=12).value

    if not category:
        continue

    rows.append({
        'category': str(category).strip() if category else '',
        'topic_en': str(topic_en).strip() if topic_en else '',
        'topic_th': str(topic_th).strip() if topic_th else '',
        'planet_en': str(planet_en).strip() if planet_en and str(planet_en) != '-' else '',
        'planet_th': str(planet_th).strip() if planet_th and str(planet_th) != '-' else '',
        'keyword': str(keyword).strip() if keyword and str(keyword) != '-' else '',
        'desc_th': str(desc_th).strip() if desc_th else '',
        'desc_en': str(desc_en).strip() if desc_en else '',
        'page': str(page).strip() if page and str(page) != '-' else '',
        'source': str(source).strip() if source and str(source) != '-' else '',
        'confidence': str(confidence).strip() if confidence and str(confidence) != '-' else '',
    })

print(f"Total parsed rows: {len(rows)}")

# Append ASTROCARTOGRAPHY_DB to database_extra.js (keeps all existing *_DB variables intact)
with open(db_extra_path, 'a', encoding='utf-8') as f:
    f.write("\n\n// Astrocartography Database (ACG) - Auto-generated from Planetary.xlsx, sheet 'Astrocartography'\n")
    f.write("// Source: Astrolocality Astrology by Martin Davis (Appendix 1 by Jeff Jawer) + supplementary notes\n")
    f.write("const ASTROCARTOGRAPHY_DB = ")
    json.dump(rows, f, ensure_ascii=False, indent=2)
    f.write(";\n")

print("Done extraction! Appended ASTROCARTOGRAPHY_DB to database_extra.js")
